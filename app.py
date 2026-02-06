import random, csv, uuid, time, os
from flask import Flask, request, render_template, jsonify, make_response
from datetime import timedelta
from collections import Counter
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class BadRequestError(Exception):
    def __init__(self, message='', success=False):
        self.message = message
        self.success = success

    def re(self):
        return jsonify({
            'success': self.success,
            'message': self.message,
        })


app = Flask(__name__,template_folder='templates')

questions = []
questions_no_correct_answer = []
players = {}


# 用于校验参赛选手身份
def verify(cookies):
    player_name = cookies.get('playerName')
    player_id = cookies.get('playerId')
    if player_name in players and players[player_name]['playerId'] == player_id:
        return True
    else:
        return False


# 选手答题页面
@app.route('/', methods=['get'])
def home():
    return render_template('player.html')


# 选手登录接口
@app.route('/login', methods=['post'])
def login():
    try:
        name = request.get_json()['playerName']
        if name in players:
            raise BadRequestError('用户名已被使用，请选择其他姓名')
        elif len(questions) <= 0:
            raise BadRequestError('请等待管理员抽取题目')

        else:
            players[name] = {
                'playerId': uuid.uuid4().hex,
                'playerName': name,
                'answers': [
                    {
                        "questionIndex": i,
                        'questionId': questions[i]['id'],
                        'question': questions[i]['question'],
                        'playerAnswer': None,
                        'correctAnswer': None,
                        'isCorrect': None,
                        'saveAnswerTimestamp': None,
                        'score': None
                    } for i in range(len(questions))],

                'loginTimestamp': time.time(),
                'timerInterval': None,
                'currentQuestionIndex': 0
            }

            response = {
                'success': True,
                'playerId': players[name]['playerId'],
                'playerName': name,
                'totalQuestions': len(questions)
            }
            response = make_response(response)
            response.set_cookie('playerId', players[name]['playerId'])
            response.set_cookie('playerName', name)
            return response
    except BadRequestError as e:
        return e.re()


# 选手加载问题接口
@app.route('/getQuestion', methods=['get'])
def get_question():
    try:
        if not verify(request.cookies):
            raise BadRequestError('身份校验不通过')
        player_name = request.cookies.get('playerName')
        current_question_index = players[player_name]['currentQuestionIndex']
        if current_question_index < 0 or current_question_index >= len(questions):
            raise BadRequestError('答题完毕或题号小于0')
        if len(questions) <= 0:
            raise BadRequestError('请等待管理员抽取题目')
        question = questions_no_correct_answer[current_question_index]
        players[player_name]['answers'][current_question_index]['startTimestamp'] = time.time()

        return jsonify({'question': question,
                        'currentQuestionIndex': current_question_index,
                        'success': True})
    except BadRequestError as e:
        return e.re()


def check_answer(question: dict, player_answer) :
    if not player_answer or (isinstance(player_answer, list) and len(player_answer) == 0):
        return False
    if question.get('type') == 'multiple':
        correct_ans = question['correctAnswer']
        return len(player_answer) == len(correct_ans) and all(ans in correct_ans for ans in player_answer)
    else:
        return player_answer == question['correctAnswer'][0]


# 选手提交答案接口
@app.route('/save-answer', methods=['post'])
def save_answer():
    try:
        data = request.get_json()
        if not verify(request.cookies):
            raise BadRequestError('用户校验错误')

        player_name = request.cookies.get('playerName')
        player = players[player_name]
        current_question_index = player['currentQuestionIndex']
        if current_question_index != data['questionIndex'] or player['answers'][current_question_index]['questionId'] != \
                data[
                    'questionId']:
            raise BadRequestError('题号校验错误')

        else:
            question = questions[current_question_index]
            answer = {
                'playerAnswer': data['answer'],
                'correctAnswer': question['correctAnswer'],
                'isCorrect': check_answer(question, data['answer']),
                'score': question['score'] if check_answer(question, data['answer']) else 0,
                'saveAnswerTimestamp': time.time(),
            }
            for key, value in answer.items():
                players[player_name]['answers'][current_question_index][key] = value

            player['currentQuestionIndex'] += 1
            return jsonify({'success': True})
    except BadRequestError as e:
        return e.re()


# 选手答题完毕后获取答题详情
@app.route('/get-answers', methods=['get'])
def get_answers():
    try:
        if verify(request.cookies):
            player_name = request.cookies.get('playerName')
            players[player_name]['timerInterval'] = str(timedelta(
                seconds=round(
                    players[player_name]['answers'][-1]['saveAnswerTimestamp'] - players[player_name]['answers'][0][
                        'startTimestamp'])))
            return jsonify({
                'success': 'true',
                'player': players[player_name]
            })
        else:
            raise BadRequestError('用户校验错误')
    except BadRequestError as e:
        return e.re()


# 大屏幕展示比赛进度页面
@app.route('/show', methods=['get'])
def show():
    return render_template('show.html')


# 获取所有选手比赛进度的接口，管理员页面也会用到这个接口
@app.route('/get_player_data', methods=['get'])
def get_player_data():
    response = []
    for playerName, playerData in players.items():
        answers = [(playerData['answers'][i]['isCorrect']) for i in
                   range(len(playerData['answers']))]
        true_count = Counter(answers)[True]
        total_elem = Counter(answers)[True] + Counter(answers)[False]
        if total_elem == 0:
            true_percent = 0
        else:
            true_percent = true_count / total_elem
        true_percent_rounded = round(true_percent, 4)
        score = sum([(playerData['answers'][i]['score']) for i in
                     range(min(playerData['currentQuestionIndex'], len(playerData['answers'])))])
        response.append({
            'name': playerData['playerName'],
            'answers': answers,
            'accuracy': true_percent_rounded,
            'score': score
        })
    return jsonify(response)


# 管理员页面
@app.route('/admin', methods=['get'])
def admin():
    return render_template('admin.html')


# 管理员从题库抽题的接口
@app.route('/load_questions', methods=['post'])
def load_questions():
    try:
        bank_url = request.get_json()['bankUrl']
        question_count = request.get_json()['questionCount']

        if not isinstance(bank_url, str) or not isinstance(question_count, int) or question_count < 1:
            raise BadRequestError('题库文件路径或抽取题量的参数错误')
        if not os.path.exists(bank_url):
            raise BadRequestError('题库文件不存在，请检查文件路径')

        wb = load_workbook(bank_url, read_only=True, data_only=True)
        ws = wb.active
        print('该文件最大行为：',ws.max_row)
        if question_count > ws.max_row - 1:
            raise BadRequestError('抽题数量不能大于题库数量')
        result_questions = []
        for row_num in random.sample(range(2, ws.max_row + 1), question_count):

            row_data = [cell.value for cell in ws[row_num]]
            question = {
                'id': str(row_data[0]),
                'type': 'single' if row_data[9] == '单选' else 'multiple' if row_data[9] == '多选' else 'true_false',
                'question': str(row_data[1]),
                'options': [str(value) for value in row_data[3:9] if value is not None],
                'correctAnswer': [str(value) for value in row_data[10:16] if value is not None],
                'score': row_data[2]
            }
            for option in question['correctAnswer']:
                if option not in question['options']:
                    raise BadRequestError('选项不在正确答案中')
            result_questions.append(question)
        return jsonify({'success': True,
                        'questions': result_questions})
    except BadRequestError as e:
        return e.re()
    except InvalidFileException as e:
        return jsonify({
            'message': '请输入正确的路径并确认题库文件是.xlsx类型',
            'success': False
        })
    finally:
        if 'wb' in locals():
            wb.close()


# 管理员保存题目的接口
@app.route('/save_questions', methods=['post'])
def save_questions():
    try:
        result_questions = request.get_json()['currentQuestions']
        if not isinstance(result_questions, list) or len(result_questions) == 0:
            raise BadRequestError('保存题库错误')
        if len(players) > 0:
            raise BadRequestError('已有选手登录，请先删除所有选手后再保存题目')
        else:
            global questions, questions_no_correct_answer
            questions = result_questions
            questions_no_correct_answer = []
            for question in questions:
                questions_no_correct_answer.append(
                    {key: question[key] for key in question.keys() & {'id', 'type', 'question', 'options', 'score'}})
            return jsonify({'success': True,
                            'questions': result_questions})
    except BadRequestError as e:
        return e.re()


# 管理员获取用户答题详情的接口
@app.route('/admin_get_answers/<player_name>', methods=['get'])
def admin_get_answers(player_name):
    try:
        if player_name not in players:
            raise BadRequestError('无此用户')
        return jsonify(
            # 'success': 'true',
            players[player_name]['answers'])
    except BadRequestError as e:
        return e.re()


# 管理员删除用户
@app.route('/delete/<player_name>', methods=['DELETE'])
def admin_delete_answers(player_name):
    try:
        if player_name not in players:
            raise BadRequestError('无此用户')
        players.pop(player_name)
        return jsonify({
            'success': True
        })
    except BadRequestError as e:
        return e.re()


# 用于生成用户答题详情
def answer_data_to_csv(answer_dict, csv_file_path):
    if not answer_dict:
        raise BadRequestError('用户答题数据为空')
    kv = {
        '选手名': 'playerName', '选手Id': 'playerId', '登录时间': 'loginTimestamp',
        '总计耗时': 'timerInterval', '问题序号': 'questionIndex', '问题Id': 'questionId',
        '问题题干': 'question', '用户答案': 'playerAnswer', '正确答案': 'correctAnswer', '是否正确': 'isCorrect',
        '得分': 'score', '获得问题时间': 'startTimestamp', '提交答案时间': 'saveAnswerTimestamp',
        'currentQuestionIndex': 'currentQuestionIndex'
    }
    fieldnames = [key for key in kv.keys()]

    with open(csv_file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()

        for player_key, player_data in answer_dict.items():
            base_info = {
                "playerId": player_data.get("playerId", ""),
                "playerName": player_data.get("playerName", ""),
                "timerInterval": player_data.get("timerInterval", ""),
                "currentQuestionIndex": player_data.get("currentQuestionIndex", "")
            }

            try:
                login_ts = player_data.get('loginTimestamp')
                base_info['loginTimestamp'] = time.strftime(
                    '%Y-%m-%d %H:%M:%S', time.localtime(login_ts)) if login_ts else ''
            except:
                base_info['loginTimestamp'] = ''

            answers = player_data.get("answers", [])

            for answer in answers:
                player_answer = answer.get('playerAnswer', '')
                row_data = base_info.copy()
                row_data.update({
                    "questionIndex": answer.get("questionIndex", "") + 1,
                    "questionId": answer.get("questionId", ""),
                    "question": answer.get("question", ""),
                    "playerAnswer": ', '.join(player_answer) if isinstance(player_answer, list) else str(
                        player_answer),
                    "isCorrect": answer.get("isCorrect", ""),
                    "score": answer.get("score", "")
                })

                correct_ans = answer.get("correctAnswer", [])
                row_data["correctAnswer"] = ", ".join(correct_ans) if isinstance(correct_ans, list) else str(
                    correct_ans)

                for ts_key in ["saveAnswerTimestamp", "startTimestamp"]:
                    ts = answer.get(ts_key)
                    row_data[ts_key] = ts if ts else ""
                    try:
                        row_data[ts_key] = time.strftime(
                            '%Y-%m-%d %H:%M:%S', time.localtime(ts)) if ts else ''
                    except:
                        row_data[ts_key] = ""
                row_data_write = {key: row_data.get(value) for key, value in kv.items()}
                writer.writerow(row_data_write)


# 用于生成问题详情
def question_to_csv(data, csv_file_path):
    if not data:
        raise BadRequestError('问题数据为空')
    kv = {
        '题目id': 'id',
        '题型': 'type',
        '题干': 'question',
        '选项': 'options',
        '正确答案': 'correctAnswer',
        '分数': 'score'
    }
    fieldnames = [key for key in kv.keys()]
    with open(csv_file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for row in data:
            row_copy = row.copy()
            for key, value in row_copy.items():
                if isinstance(value, list):
                    row_copy[key] = ', '.join(value)
            row_data = {key: row_copy.get(value) for key, value in kv.items()}
            writer.writerow(row_data)


# 管理员保存用户答题详情和问题详情日志
@app.route('/save_log', methods=['get'])
def save_log():
    try:
        now_time = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime())
        question_to_csv(questions, f'log/问题数据{now_time}.csv')
        answer_data_to_csv(players, f'log/答题数据{now_time}.csv')

        return jsonify({'success': 'true'})
    except BadRequestError as e:
        return e.re()
    except Exception as e:
        print('save_log error', e)
        return jsonify({'success': 'false'})


if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)
